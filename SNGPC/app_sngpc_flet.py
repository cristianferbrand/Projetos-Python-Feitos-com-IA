# app_sngpc_flet_v3.py
# Requisitos:
#   pip install flet openpyxl
#
# Executar:
#   python app_sngpc_flet_v3.py

import sys
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import xml.etree.ElementTree as ET

import flet as ft
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------
# Compat shim (ft.Colors / ft.colors, ft.Icons / ft.icons)
# ---------------------------
try:
    C = ft.Colors
except AttributeError:
    C = ft.colors

try:
    I = ft.Icons
except AttributeError:
    I = ft.icons


def op(alpha: float, color):
    """Opacity helper compatible across flet versions."""
    if hasattr(C, "with_opacity"):
        return C.with_opacity(alpha, color)
    try:
        return ft.colors.with_opacity(alpha, color)
    except Exception:
        return color


# ---------------------------
# Constantes / colunas
# ---------------------------
NS = {"s": "urn:sngpc-schema"}

ENTRADAS_COLS = [
    "cnpjEmissor",
    "cpfTransmissor",
    "dataInicio",
    "dataFim",
    "numeroNotaFiscal",
    "tipoOperacaoNotaFiscal",
    "dataNotaFiscal",
    "cnpjOrigem",
    "cnpjDestino",
    "classeTerapeutica",
    "registroMSMedicamento",
    "numeroLoteMedicamento",
    "quantidadeMedicamento",
    "unidadeMedidaMedicamento",
    "dataRecebimentoMedicamento",
]

SAIDAS_COLS = [
    "cnpjEmissor",
    "cpfTransmissor",
    "dataInicio",
    "dataFim",
    "tipoReceituarioMedicamento",
    "numeroNotificacaoMedicamento",
    "dataPrescricaoMedicamento",
    "nomePrescritor",
    "numeroRegistroProfissional",
    "conselhoProfissional",
    "UFConselho",
    "usoMedicamento",
    "nomeComprador",
    "tipoDocumento",
    "numeroDocumento",
    "orgaoExpedidor",
    "UFEmissaoDocumento",
    "usoProlongado",
    "registroMSMedicamento",
    "numeroLoteMedicamento",
    "quantidadeMedicamento",
    "unidadeMedidaMedicamento",
    "dataVendaMedicamento",
]


# ---------------------------
# Utilidades
# ---------------------------
def _text(elem, path: str) -> str:
    if elem is None:
        return ""
    e = elem.find(path, NS)
    if e is None or e.text is None:
        return ""
    return e.text.strip()


def _parse_date(s: str):
    s = (s or "").strip()
    if not s:
        return ""
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return s


def _parse_decimal(s: str) -> Decimal:
    s = (s or "").strip()
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _fmt_date(v) -> str:
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    return str(v) if v is not None else ""


def _fmt_num(v) -> str:
    if isinstance(v, Decimal):
        s = format(v, "f")
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s
    return str(v) if v is not None else ""


def get_base_dir() -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def safe_int(v, default=200):
    try:
        return int(v)
    except Exception:
        return default


def clamp(n, min_n, max_n):
    return max(min_n, min(n, max_n))


# ---------------------------
# Parse SNGPC
# ---------------------------
def build_resumo(header: dict, entradas: list[dict], saidas: list[dict]) -> dict:
    mov_entr = len(entradas)
    mov_sai = len(saidas)

    tot_entr = sum((r.get("quantidadeMedicamento") or Decimal("0")) for r in entradas)
    tot_sai = sum((r.get("quantidadeMedicamento") or Decimal("0")) for r in saidas)

    key_lote = set()
    key_reg = set()

    for r in entradas:
        k = (
            str(r.get("registroMSMedicamento", "")).strip(),
            str(r.get("numeroLoteMedicamento", "")).strip(),
            str(r.get("unidadeMedidaMedicamento", "")).strip(),
        )
        if any(k):
            key_lote.add(k)
        reg = str(r.get("registroMSMedicamento", "")).strip()
        if reg:
            key_reg.add(reg)

    for r in saidas:
        k = (
            str(r.get("registroMSMedicamento", "")).strip(),
            str(r.get("numeroLoteMedicamento", "")).strip(),
            str(r.get("unidadeMedidaMedicamento", "")).strip(),
        )
        if any(k):
            key_lote.add(k)
        reg = str(r.get("registroMSMedicamento", "")).strip()
        if reg:
            key_reg.add(reg)

    periodo = f"{_fmt_date(header.get('dataInicio'))} a {_fmt_date(header.get('dataFim'))}"

    return {
        "CNPJ Emissor": header.get("cnpjEmissor", ""),
        "CPF Transmissor": header.get("cpfTransmissor", ""),
        "Período": periodo,
        "Movimentos de Entrada (linhas)": mov_entr,
        "Total Entrada (quantidade)": tot_entr,
        "Movimentos de Saída (linhas)": mov_sai,
        "Total Saída (quantidade)": tot_sai,
        "Produtos distintos (Registro+Lote+Unidade)": len(key_lote),
        "Produtos distintos (somente Registro)": len(key_reg),
    }


def parse_sngpc_xml(xml_path: str):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    cab = root.find("s:cabecalho", NS)
    header = {
        "cnpjEmissor": _text(cab, "s:cnpjEmissor"),
        "cpfTransmissor": _text(cab, "s:cpfTransmissor"),
        "dataInicio": _parse_date(_text(cab, "s:dataInicio")),
        "dataFim": _parse_date(_text(cab, "s:dataFim")),
    }

    corpo = root.find("s:corpo", NS)
    meds = None
    if corpo is not None:
        meds = corpo.find("s:medicamentos", NS)

    entradas = []
    saidas = []

    if meds is not None:
        for ent in meds.findall("s:entradaMedicamentos", NS):
            nota = ent.find("s:notaFiscalEntradaMedicamento", NS)
            med_ent = ent.find("s:medicamentoEntrada", NS)
            row = {
                **header,
                "numeroNotaFiscal": _text(nota, "s:numeroNotaFiscal"),
                "tipoOperacaoNotaFiscal": _text(nota, "s:tipoOperacaoNotaFiscal"),
                "dataNotaFiscal": _parse_date(_text(nota, "s:dataNotaFiscal")),
                "cnpjOrigem": _text(nota, "s:cnpjOrigem"),
                "cnpjDestino": _text(nota, "s:cnpjDestino"),
                "classeTerapeutica": _text(med_ent, "s:classeTerapeutica"),
                "registroMSMedicamento": _text(med_ent, "s:registroMSMedicamento"),
                "numeroLoteMedicamento": _text(med_ent, "s:numeroLoteMedicamento"),
                "quantidadeMedicamento": _parse_decimal(_text(med_ent, "s:quantidadeMedicamento")),
                "unidadeMedidaMedicamento": _text(med_ent, "s:unidadeMedidaMedicamento"),
                "dataRecebimentoMedicamento": _parse_date(_text(ent, "s:dataRecebimentoMedicamento")),
            }
            entradas.append(row)

        for sai in meds.findall("s:saidaMedicamentoVendaAoConsumidor", NS):
            pres = sai.find("s:prescritorMedicamento", NS)
            comp = sai.find("s:compradorMedicamento", NS)
            med_v = sai.find("s:medicamentoVenda", NS)

            uso_prol = _text(med_v, "s:usoProlongado")
            if uso_prol.strip().lower() in ("true", "1", "sim", "s"):
                uso_prol_fmt = "Sim"
            elif uso_prol:
                uso_prol_fmt = "Não"
            else:
                uso_prol_fmt = ""

            row = {
                **header,
                "tipoReceituarioMedicamento": _text(sai, "s:tipoReceituarioMedicamento"),
                "numeroNotificacaoMedicamento": _text(sai, "s:numeroNotificacaoMedicamento"),
                "dataPrescricaoMedicamento": _parse_date(_text(sai, "s:dataPrescricaoMedicamento")),
                "nomePrescritor": _text(pres, "s:nomePrescritor"),
                "numeroRegistroProfissional": _text(pres, "s:numeroRegistroProfissional"),
                "conselhoProfissional": _text(pres, "s:conselhoProfissional"),
                "UFConselho": _text(pres, "s:UFConselho"),
                "usoMedicamento": _text(sai, "s:usoMedicamento"),
                "nomeComprador": _text(comp, "s:nomeComprador"),
                "tipoDocumento": _text(comp, "s:tipoDocumento"),
                "numeroDocumento": _text(comp, "s:numeroDocumento"),
                "orgaoExpedidor": _text(comp, "s:orgaoExpedidor"),
                "UFEmissaoDocumento": _text(comp, "s:UFEmissaoDocumento"),
                "usoProlongado": uso_prol_fmt,
                "registroMSMedicamento": _text(med_v, "s:registroMSMedicamento"),
                "numeroLoteMedicamento": _text(med_v, "s:numeroLoteMedicamento"),
                "quantidadeMedicamento": _parse_decimal(_text(med_v, "s:quantidadeMedicamento")),
                "unidadeMedidaMedicamento": _text(med_v, "s:unidadeMedidaMedicamento"),
                "dataVendaMedicamento": _parse_date(_text(sai, "s:dataVendaMedicamento")),
            }
            saidas.append(row)

    resumo = build_resumo(header, entradas, saidas)
    return header, entradas, saidas, resumo


# ---------------------------
# Export XLSX (4 abas)
# ---------------------------
def export_xlsx_4abas(output_path: str, header: dict, entradas: list[dict], saidas: list[dict], resumo: dict):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=False)

    def _apply_header_row(ws, col_count: int):
        for col_idx in range(1, col_count + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = hdr_align
            c.border = border

    def _style_body(ws, max_row: int, max_col: int):
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = cell_align

    def _auto_width(ws, max_col: int, max_row: int):
        for i in range(1, max_col + 1):
            max_len = 10
            hv = ws.cell(row=1, column=i).value
            if hv is not None:
                max_len = max(max_len, len(str(hv)))
            for r in range(2, min(max_row, 300) + 1):
                val = ws.cell(row=r, column=i).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 45)

    def _add_table(ws, name: str):
        max_row = ws.max_row
        max_col = ws.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    ws_h = wb.create_sheet("Cabecalho")
    ws_h.append(["campo", "valor"])
    for k in ["cnpjEmissor", "cpfTransmissor", "dataInicio", "dataFim"]:
        v = header.get(k, "")
        if isinstance(v, date):
            v = v.strftime("%d/%m/%Y")
        ws_h.append([k, v])

    _apply_header_row(ws_h, 2)
    _style_body(ws_h, ws_h.max_row, ws_h.max_column)
    ws_h.freeze_panes = "A2"
    ws_h.sheet_view.showGridLines = False
    ws_h.auto_filter.ref = f"A1:B{ws_h.max_row}"
    ws_h.column_dimensions["A"].width = 22
    ws_h.column_dimensions["B"].width = 28

    ws_e = wb.create_sheet("Entradas")
    ws_e.append(ENTRADAS_COLS)
    for r in entradas:
        row = []
        for col in ENTRADAS_COLS:
            v = r.get(col, "")
            if isinstance(v, date):
                row.append(v.strftime("%d/%m/%Y"))
            elif isinstance(v, Decimal):
                row.append(float(v))
            else:
                row.append(v)
        ws_e.append(row)

    _apply_header_row(ws_e, len(ENTRADAS_COLS))
    _style_body(ws_e, ws_e.max_row, ws_e.max_column)
    ws_e.freeze_panes = "A2"
    ws_e.sheet_view.showGridLines = False
    ws_e.auto_filter.ref = f"A1:{get_column_letter(ws_e.max_column)}{ws_e.max_row}"
    _auto_width(ws_e, ws_e.max_column, ws_e.max_row)
    _add_table(ws_e, "Entradas")

    ws_s = wb.create_sheet("Saidas")
    ws_s.append(SAIDAS_COLS)
    for r in saidas:
        row = []
        for col in SAIDAS_COLS:
            v = r.get(col, "")
            if isinstance(v, date):
                row.append(v.strftime("%d/%m/%Y"))
            elif isinstance(v, Decimal):
                row.append(float(v))
            else:
                row.append(v)
        ws_s.append(row)

    _apply_header_row(ws_s, len(SAIDAS_COLS))
    _style_body(ws_s, ws_s.max_row, ws_s.max_column)
    ws_s.freeze_panes = "A2"
    ws_s.sheet_view.showGridLines = False
    ws_s.auto_filter.ref = f"A1:{get_column_letter(ws_s.max_column)}{ws_s.max_row}"
    _auto_width(ws_s, ws_s.max_column, ws_s.max_row)
    _add_table(ws_s, "Saidas")

    ws_r = wb.create_sheet("Resumo")
    ws_r.append(["Indicador", "Valor"])
    _apply_header_row(ws_r, 2)

    for k, v in resumo.items():
        if isinstance(v, Decimal):
            v = float(v)
        ws_r.append([k, v])

    _style_body(ws_r, ws_r.max_row, ws_r.max_column)
    ws_r.freeze_panes = "A2"
    ws_r.sheet_view.showGridLines = False
    ws_r.auto_filter.ref = f"A1:B{ws_r.max_row}"
    ws_r.column_dimensions["A"].width = 46
    ws_r.column_dimensions["B"].width = 30

    wb.save(output_path)


# ---------------------------
# UI Helpers
# ---------------------------
def paginate(items: list, offset: int, limit: int):
    total = len(items)
    if total == 0:
        return [], 0, 0, 0
    limit = max(1, limit)
    offset = clamp(offset, 0, max(0, total - 1))
    start = offset
    end = min(offset + limit, total)
    return items[start:end], start, end, total


def main(page: ft.Page):
    page.title = "SNGPC - Análise (Cabeçalho / Entradas / Saídas / Resumo)"
    page.window_width = 1200
    page.window_height = 760
    page.theme_mode = ft.ThemeMode.LIGHT
    page.scroll = ft.ScrollMode.AUTO

    state = {
        "xml_path": None,
        "header": {},
        "entradas": [],
        "saidas": [],
        "resumo": {},
        "entradas_f": [],
        "saidas_f": [],
        "filtro_entr": "",
        "filtro_sai": "",
        "entr_offset": 0,
        "sai_offset": 0,
        "entr_limit": 200,
        "sai_limit": 200,
    }

    def notify(msg: str, error: bool = False):
        page.snack_bar = ft.SnackBar(
            content=ft.Text(msg),
            bgcolor=(op(0.90, C.RED) if error else op(0.90, C.GREEN)),
        )
        page.snack_bar.open = True
        page.update()

    def set_loading(is_loading: bool, text: str = ""):
        loading.visible = is_loading
        loading_text.value = text
        btn_open.disabled = is_loading
        btn_export.disabled = is_loading or (state["xml_path"] is None)
        btn_export.visible = state["xml_path"] is not None
        page.update()

    pick_xml = ft.FilePicker()
    save_xlsx = ft.FilePicker()
    page.overlay.extend([pick_xml, save_xlsx])

    header_kv = ft.Column(spacing=8)

    # DataTables precisam nascer com ao menos 1 coluna
    entradas_table = ft.DataTable(
        columns=[ft.DataColumn(ft.Text("Abra um XML primeiro"))],
        rows=[],
        heading_row_color=op(0.08, C.BLUE),
        data_row_min_height=36,
        data_row_max_height=44,
        column_spacing=14,
        horizontal_lines=ft.BorderSide(1, op(0.2, C.GREY)),
    )
    saidas_table = ft.DataTable(
        columns=[ft.DataColumn(ft.Text("Abra um XML primeiro"))],
        rows=[],
        heading_row_color=op(0.08, C.BLUE),
        data_row_min_height=36,
        data_row_max_height=44,
        column_spacing=14,
        horizontal_lines=ft.BorderSide(1, op(0.2, C.GREY)),
    )

    resumo_cards = ft.Column(spacing=10)

    txt_arquivo = ft.Text(value="Nenhum XML carregado.", color=C.GREY)

    filtro_entr = ft.TextField(
        label="Filtro (Entradas)",
        hint_text="Digite parte do Registro MS, Lote, Nota, CNPJ...",
        prefix_icon=I.SEARCH,
        expand=True,
    )
    filtro_sai = ft.TextField(
        label="Filtro (Saídas)",
        hint_text="Digite parte do Registro MS, Lote, Comprador, Documento...",
        prefix_icon=I.SEARCH,
        expand=True,
    )

    dd_entr_limit = ft.Dropdown(
        label="Linhas por página",
        value="200",
        width=160,
        options=[ft.dropdown.Option(x) for x in ["50", "100", "200", "500", "1000"]],
    )
    dd_sai_limit = ft.Dropdown(
        label="Linhas por página",
        value="200",
        width=160,
        options=[ft.dropdown.Option(x) for x in ["50", "100", "200", "500", "1000"]],
    )

    lbl_entr_count = ft.Text("Entradas: 0", color=C.GREY)
    lbl_sai_count = ft.Text("Saídas: 0", color=C.GREY)
    lbl_entr_page = ft.Text("", color=C.GREY)
    lbl_sai_page = ft.Text("", color=C.GREY)

    btn_entr_prev = ft.IconButton(icon=I.CHEVRON_LEFT, tooltip="Página anterior")
    btn_entr_next = ft.IconButton(icon=I.CHEVRON_RIGHT, tooltip="Próxima página")
    btn_sai_prev = ft.IconButton(icon=I.CHEVRON_LEFT, tooltip="Página anterior")
    btn_sai_next = ft.IconButton(icon=I.CHEVRON_RIGHT, tooltip="Próxima página")

    loading_text = ft.Text("", color=C.GREY)
    loading = ft.Row(
        [
            ft.ProgressRing(width=18, height=18, stroke_width=2),
            loading_text,
        ],
        visible=False,
        spacing=10,
    )

    def build_header_view():
        header_kv.controls.clear()
        if not state["header"]:
            header_kv.controls.append(ft.Text("Sem dados. Abra um XML.", color=C.GREY))
            return

        items = [
            ("CNPJ Emissor", state["header"].get("cnpjEmissor", "")),
            ("CPF Transmissor", state["header"].get("cpfTransmissor", "")),
            ("Data Início", _fmt_date(state["header"].get("dataInicio", ""))),
            ("Data Fim", _fmt_date(state["header"].get("dataFim", ""))),
        ]

        for k, v in items:
            header_kv.controls.append(
                ft.Container(
                    content=ft.Row(
                        [
                            ft.Text(k, weight=ft.FontWeight.BOLD, width=160),
                            ft.Text(str(v)),
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    ),
                    padding=12,
                    bgcolor=op(0.03, C.BLACK),
                    border_radius=10,
                )
            )

    def build_datatable(table: ft.DataTable, cols: list[str], rows: list[dict]):
        if not cols:
            cols = ["Sem colunas"]

        table.columns = [ft.DataColumn(ft.Text(c)) for c in cols]

        dt_rows = []
        for r in rows:
            cells = []
            for c in cols:
                v = r.get(c, "")
                if isinstance(v, date):
                    v_str = v.strftime("%d/%m/%Y")
                elif isinstance(v, Decimal):
                    v_str = _fmt_num(v)
                else:
                    v_str = str(v) if v is not None else ""

                cells.append(
                    ft.DataCell(
                        ft.Text(
                            v_str,
                            no_wrap=True,
                            overflow=ft.TextOverflow.ELLIPSIS,
                            selectable=True,
                        )
                    )
                )
            dt_rows.append(ft.DataRow(cells=cells))

        table.rows = dt_rows

    def build_resumo_view():
        resumo_cards.controls.clear()
        if not state["resumo"]:
            resumo_cards.controls.append(ft.Text("Sem dados. Abra um XML.", color=C.GREY))
            return

        kpi_items = [
            ("Mov. Entradas", state["resumo"].get("Movimentos de Entrada (linhas)", 0)),
            ("Total Entradas", _fmt_num(state["resumo"].get("Total Entrada (quantidade)", Decimal("0")))),
            ("Mov. Saídas", state["resumo"].get("Movimentos de Saída (linhas)", 0)),
            ("Total Saídas", _fmt_num(state["resumo"].get("Total Saída (quantidade)", Decimal("0")))),
        ]
        kpis = ft.Row(
            controls=[
                ft.Container(
                    padding=12,
                    border_radius=12,
                    bgcolor=op(0.03, C.BLACK),
                    expand=True,
                    content=ft.Column(
                        [
                            ft.Text(lbl, color=C.GREY, size=12),
                            ft.Text(str(val), weight=ft.FontWeight.BOLD, size=18),
                        ],
                        spacing=4,
                    ),
                )
                for lbl, val in kpi_items
            ],
            spacing=10,
        )
        resumo_cards.controls.append(kpis)

        others = [
            ("Produtos distintos (Registro+Lote+Unidade)", state["resumo"].get("Produtos distintos (Registro+Lote+Unidade)", 0)),
            ("Produtos distintos (somente Registro)", state["resumo"].get("Produtos distintos (somente Registro)", 0)),
            ("Período", state["resumo"].get("Período", "")),
        ]

        for k, v in others:
            resumo_cards.controls.append(
                ft.Container(
                    content=ft.Row(
                        [
                            ft.Text(k, weight=ft.FontWeight.BOLD),
                            ft.Text(str(v)),
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    ),
                    padding=12,
                    bgcolor=op(0.03, C.BLACK),
                    border_radius=10,
                )
            )

    def match_any(row: dict, q: str) -> bool:
        if not q:
            return True
        q = q.lower()
        for v in row.values():
            if v is None:
                continue
            if isinstance(v, date):
                s = _fmt_date(v)
            elif isinstance(v, Decimal):
                s = _fmt_num(v)
            else:
                s = str(v)
            if q in s.lower():
                return True
        return False

    def refresh_entr_table():
        items = state["entradas_f"]
        limit = state["entr_limit"]
        offset = state["entr_offset"]

        page_items, start, end, total = paginate(items, offset, limit)

        if total > 0 and start >= total:
            state["entr_offset"] = max(0, (max(0, total - 1) // limit) * limit)
            page_items, start, end, total = paginate(items, state["entr_offset"], limit)

        build_datatable(entradas_table, ENTRADAS_COLS, page_items)

        if total == 0:
            lbl_entr_page.value = "Mostrando 0 de 0"
        else:
            page_num = (start // limit) + 1
            page_total = ((total - 1) // limit) + 1
            lbl_entr_page.value = f"Mostrando {start+1}-{end} de {total} | Página {page_num}/{page_total}"

        btn_entr_prev.disabled = (total == 0 or start == 0)
        btn_entr_next.disabled = (total == 0 or end >= total)

    def refresh_sai_table():
        items = state["saidas_f"]
        limit = state["sai_limit"]
        offset = state["sai_offset"]

        page_items, start, end, total = paginate(items, offset, limit)

        if total > 0 and start >= total:
            state["sai_offset"] = max(0, (max(0, total - 1) // limit) * limit)
            page_items, start, end, total = paginate(items, state["sai_offset"], limit)

        build_datatable(saidas_table, SAIDAS_COLS, page_items)

        if total == 0:
            lbl_sai_page.value = "Mostrando 0 de 0"
        else:
            page_num = (start // limit) + 1
            page_total = ((total - 1) // limit) + 1
            lbl_sai_page.value = f"Mostrando {start+1}-{end} de {total} | Página {page_num}/{page_total}"

        btn_sai_prev.disabled = (total == 0 or start == 0)
        btn_sai_next.disabled = (total == 0 or end >= total)

    def apply_filters():
        state["filtro_entr"] = (filtro_entr.value or "").strip().lower()
        state["filtro_sai"] = (filtro_sai.value or "").strip().lower()

        entr_f = [r for r in state["entradas"] if match_any(r, state["filtro_entr"])]
        sai_f = [r for r in state["saidas"] if match_any(r, state["filtro_sai"])]

        state["entradas_f"] = entr_f
        state["saidas_f"] = sai_f

        lbl_entr_count.value = f"Entradas: {len(entr_f)} (de {len(state['entradas'])})"
        lbl_sai_count.value = f"Saídas: {len(sai_f)} (de {len(state['saidas'])})"

        state["entr_offset"] = 0
        state["sai_offset"] = 0

        refresh_entr_table()
        refresh_sai_table()
        page.update()

    def on_filtro_entr_change(e):
        apply_filters()

    def on_filtro_sai_change(e):
        apply_filters()

    filtro_entr.on_change = on_filtro_entr_change
    filtro_sai.on_change = on_filtro_sai_change

    def on_entr_limit_change(e):
        state["entr_limit"] = safe_int(dd_entr_limit.value, 200)
        state["entr_offset"] = 0
        refresh_entr_table()
        page.update()

    def on_sai_limit_change(e):
        state["sai_limit"] = safe_int(dd_sai_limit.value, 200)
        state["sai_offset"] = 0
        refresh_sai_table()
        page.update()

    dd_entr_limit.on_change = on_entr_limit_change
    dd_sai_limit.on_change = on_sai_limit_change

    def on_entr_prev(_):
        state["entr_offset"] = max(0, state["entr_offset"] - state["entr_limit"])
        refresh_entr_table()
        page.update()

    def on_entr_next(_):
        state["entr_offset"] = state["entr_offset"] + state["entr_limit"]
        refresh_entr_table()
        page.update()

    def on_sai_prev(_):
        state["sai_offset"] = max(0, state["sai_offset"] - state["sai_limit"])
        refresh_sai_table()
        page.update()

    def on_sai_next(_):
        state["sai_offset"] = state["sai_offset"] + state["sai_limit"]
        refresh_sai_table()
        page.update()

    btn_entr_prev.on_click = on_entr_prev
    btn_entr_next.on_click = on_entr_next
    btn_sai_prev.on_click = on_sai_prev
    btn_sai_next.on_click = on_sai_next

    def on_pick_xml_result(e: ft.FilePickerResultEvent):
        if not e.files:
            return
        f = e.files[0]
        if not f.path:
            notify("Este modo exige caminho do arquivo (desktop).", error=True)
            return

        xml_path = f.path
        set_loading(True, "Lendo e processando XML...")

        try:
            header, entradas, saidas, resumo = parse_sngpc_xml(xml_path)

            state["xml_path"] = xml_path
            state["header"] = header
            state["entradas"] = entradas
            state["saidas"] = saidas
            state["resumo"] = resumo

            txt_arquivo.value = f"Carregado: {Path(xml_path).name} | Entradas: {len(entradas)} | Saídas: {len(saidas)}"

            build_header_view()
            build_resumo_view()

            filtro_entr.value = ""
            filtro_sai.value = ""
            dd_entr_limit.value = "200"
            dd_sai_limit.value = "200"
            state["entr_limit"] = 200
            state["sai_limit"] = 200
            state["entr_offset"] = 0
            state["sai_offset"] = 0

            apply_filters()
            notify("XML carregado com sucesso.")
        except Exception as ex:
            notify(f"Falha ao processar XML: {ex}", error=True)
        finally:
            set_loading(False, "")

    def on_save_xlsx_result(e: ft.FilePickerResultEvent):
        if not e.path:
            return
        out_path = e.path
        set_loading(True, "Exportando XLSX...")

        try:
            export_xlsx_4abas(out_path, state["header"], state["entradas"], state["saidas"], state["resumo"])
            notify(f"Exportado: {out_path}")
        except Exception as ex:
            notify(f"Falha ao exportar XLSX: {ex}", error=True)
        finally:
            set_loading(False, "")

    pick_xml.on_result = on_pick_xml_result
    save_xlsx.on_result = on_save_xlsx_result

    def open_xml(_):
        pick_xml.pick_files(allow_multiple=False, allowed_extensions=["xml"])

    def export_xlsx(_):
        if not state["xml_path"]:
            notify("Abra um XML antes de exportar.", error=True)
            return
        suggested = f"SNGPC_Analise_4Abas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_xlsx.save_file(file_name=suggested, allowed_extensions=["xlsx"])

    btn_open = ft.ElevatedButton("Abrir XML", icon=I.FOLDER_OPEN, on_click=open_xml)
    btn_export = ft.ElevatedButton("Exportar XLSX", icon=I.DOWNLOAD, on_click=export_xlsx, disabled=True, visible=False)

    tab_cabecalho = ft.Container(
        content=ft.Column(
            [
                ft.Text("Cabeçalho", size=18, weight=ft.FontWeight.BOLD),
                header_kv,
            ],
            spacing=12,
        ),
        padding=12,
    )

    # ---------------------------
    # AQUI ESTÁ O AJUSTE:
    # - Row(scroll=ALWAYS) para rolagem horizontal
    # - Column(scroll=AUTO) para rolagem vertical da área, se precisar
    # ---------------------------
    entradas_grid_area = ft.Container(
        content=ft.Column(
            [
                ft.Row(
                    [entradas_table],
                    scroll=ft.ScrollMode.ALWAYS,  # horizontal sempre
                ),
            ],
            scroll=ft.ScrollMode.AUTO,  # vertical se a área ficar menor
            expand=True,
        ),
        expand=True,
    )

    saidas_grid_area = ft.Container(
        content=ft.Column(
            [
                ft.Row(
                    [saidas_table],
                    scroll=ft.ScrollMode.ALWAYS,
                ),
            ],
            scroll=ft.ScrollMode.AUTO,
            expand=True,
        ),
        expand=True,
    )

    tab_entradas = ft.Container(
        content=ft.Column(
            [
                ft.Row(
                    [ft.Text("Entradas", size=18, weight=ft.FontWeight.BOLD), lbl_entr_count],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                ),
                ft.Row([filtro_entr, dd_entr_limit, btn_entr_prev, btn_entr_next], spacing=10),
                lbl_entr_page,
                entradas_grid_area,
            ],
            spacing=10,
            expand=True,
        ),
        padding=12,
        expand=True,
    )

    tab_saidas = ft.Container(
        content=ft.Column(
            [
                ft.Row(
                    [ft.Text("Saídas", size=18, weight=ft.FontWeight.BOLD), lbl_sai_count],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                ),
                ft.Row([filtro_sai, dd_sai_limit, btn_sai_prev, btn_sai_next], spacing=10),
                lbl_sai_page,
                saidas_grid_area,
            ],
            spacing=10,
            expand=True,
        ),
        padding=12,
        expand=True,
    )

    tab_resumo = ft.Container(
        content=ft.Column(
            [
                ft.Text("Resumo", size=18, weight=ft.FontWeight.BOLD),
                resumo_cards,
            ],
            spacing=12,
        ),
        padding=12,
    )

    tabs = ft.Tabs(
        selected_index=0,
        animation_duration=200,
        tabs=[
            ft.Tab(text="Cabeçalho", content=tab_cabecalho),
            ft.Tab(text="Entradas", content=tab_entradas),
            ft.Tab(text="Saídas", content=tab_saidas),
            ft.Tab(text="Resumo", content=tab_resumo),
        ],
        expand=True,
    )

    page.appbar = ft.AppBar(
        title=ft.Text("SNGPC - Leitor e Análise (v3)"),
        center_title=False,
        bgcolor=op(0.08, C.BLUE),
        actions=[
            ft.Container(content=btn_open, padding=ft.padding.only(right=8)),
            ft.Container(content=btn_export, padding=ft.padding.only(right=12)),
        ],
    )

    page.add(
        ft.Container(
            content=ft.Column(
                [
                    ft.Row([ft.Icon(I.DESCRIPTION, color=C.BLUE), txt_arquivo], spacing=10),
                    loading,
                    tabs,
                ],
                spacing=10,
                expand=True,
            ),
            padding=10,
            expand=True,
        )
    )

    def refresh_entr_table():
        items = state["entradas_f"]
        limit = state["entr_limit"]
        offset = state["entr_offset"]

        page_items, start, end, total = paginate(items, offset, limit)

        if total > 0 and start >= total:
            state["entr_offset"] = max(0, (max(0, total - 1) // limit) * limit)
            page_items, start, end, total = paginate(items, state["entr_offset"], limit)

        build_datatable(entradas_table, ENTRADAS_COLS, page_items)

        if total == 0:
            lbl_entr_page.value = "Mostrando 0 de 0"
        else:
            page_num = (start // limit) + 1
            page_total = ((total - 1) // limit) + 1
            lbl_entr_page.value = f"Mostrando {start+1}-{end} de {total} | Página {page_num}/{page_total}"

        btn_entr_prev.disabled = (total == 0 or start == 0)
        btn_entr_next.disabled = (total == 0 or end >= total)

    def refresh_sai_table():
        items = state["saidas_f"]
        limit = state["sai_limit"]
        offset = state["sai_offset"]

        page_items, start, end, total = paginate(items, offset, limit)

        if total > 0 and start >= total:
            state["sai_offset"] = max(0, (max(0, total - 1) // limit) * limit)
            page_items, start, end, total = paginate(items, state["sai_offset"], limit)

        build_datatable(saidas_table, SAIDAS_COLS, page_items)

        if total == 0:
            lbl_sai_page.value = "Mostrando 0 de 0"
        else:
            page_num = (start // limit) + 1
            page_total = ((total - 1) // limit) + 1
            lbl_sai_page.value = f"Mostrando {start+1}-{end} de {total} | Página {page_num}/{page_total}"

        btn_sai_prev.disabled = (total == 0 or start == 0)
        btn_sai_next.disabled = (total == 0 or end >= total)

    def apply_filters():
        state["filtro_entr"] = (filtro_entr.value or "").strip().lower()
        state["filtro_sai"] = (filtro_sai.value or "").strip().lower()

        entr_f = [r for r in state["entradas"] if match_any(r, state["filtro_entr"])]
        sai_f = [r for r in state["saidas"] if match_any(r, state["filtro_sai"])]

        state["entradas_f"] = entr_f
        state["saidas_f"] = sai_f

        lbl_entr_count.value = f"Entradas: {len(entr_f)} (de {len(state['entradas'])})"
        lbl_sai_count.value = f"Saídas: {len(sai_f)} (de {len(state['saidas'])})"

        state["entr_offset"] = 0
        state["sai_offset"] = 0

        refresh_entr_table()
        refresh_sai_table()
        page.update()

    def on_filtro_entr_change(e):
        apply_filters()

    def on_filtro_sai_change(e):
        apply_filters()

    filtro_entr.on_change = on_filtro_entr_change
    filtro_sai.on_change = on_filtro_sai_change

    def on_entr_limit_change(e):
        state["entr_limit"] = safe_int(dd_entr_limit.value, 200)
        state["entr_offset"] = 0
        refresh_entr_table()
        page.update()

    def on_sai_limit_change(e):
        state["sai_limit"] = safe_int(dd_sai_limit.value, 200)
        state["sai_offset"] = 0
        refresh_sai_table()
        page.update()

    dd_entr_limit.on_change = on_entr_limit_change
    dd_sai_limit.on_change = on_sai_limit_change

    def on_entr_prev(_):
        state["entr_offset"] = max(0, state["entr_offset"] - state["entr_limit"])
        refresh_entr_table()
        page.update()

    def on_entr_next(_):
        state["entr_offset"] = state["entr_offset"] + state["entr_limit"]
        refresh_entr_table()
        page.update()

    def on_sai_prev(_):
        state["sai_offset"] = max(0, state["sai_offset"] - state["sai_limit"])
        refresh_sai_table()
        page.update()

    def on_sai_next(_):
        state["sai_offset"] = state["sai_offset"] + state["sai_limit"]
        refresh_sai_table()
        page.update()

    btn_entr_prev.on_click = on_entr_prev
    btn_entr_next.on_click = on_entr_next
    btn_sai_prev.on_click = on_sai_prev
    btn_sai_next.on_click = on_sai_next

    pick_xml.on_result = on_pick_xml_result
    save_xlsx.on_result = on_save_xlsx_result

    # Estado inicial
    build_header_view()
    build_resumo_view()
    refresh_entr_table()
    refresh_sai_table()
    page.update()


if __name__ == "__main__":
    ft.app(target=main)