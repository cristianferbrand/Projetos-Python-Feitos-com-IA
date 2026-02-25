import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import camelot
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# Função para remover quebras de linha dentro das células
def limpar_quebras_de_linha(texto):
    return texto.replace("\n", " ") if isinstance(texto, str) else texto

# Função para padronizar o layout das tabelas
def padronizar_layout(df, colunas_padrao):
    # Verifica se o DataFrame tem todas as colunas necessárias
    for coluna in colunas_padrao:
        if coluna not in df.columns:
            df[coluna] = ""  # Adiciona a coluna ausente com valores em branco
    # Reordena as colunas para garantir o layout
    return df[colunas_padrao]

# Função para carregar PDF e salvar como Excel
def carregar_pdf():
    # Solicita ao usuário o arquivo PDF
    pdf_path = filedialog.askopenfilename(
        title="Selecione o arquivo PDF",
        filetypes=(("PDF Files", "*.pdf"), ("All Files", "*.*"))
    )
    if pdf_path:
        try:
            # Lê todas as tabelas do PDF com o layout padrão para todas as páginas
            tables = camelot.read_pdf(pdf_path, pages="all", flavor="stream", strip_text="\n")

            # Estrutura esperada das colunas
            colunas = [
                "Linha", "Posição", "Campo", "Mensagem", "Registro",
                "Conteúdo do Registro", "Conteúdo do Campo", "Valor Esperado"
            ]
            dados_completos = []

            # Processa cada tabela e verifica se é uma tabela de AVISOS
            for table in tables:
                df = table.df
                # Verifica se o título "AVISOS" está na primeira coluna e em alguma das primeiras linhas
                aviso_index = None
                for i, row in df.iterrows():
                    if "AVISOS" in row[0]:  # Verifica se "AVISOS" está na primeira coluna
                        aviso_index = i + 1  # Define o índice da linha logo abaixo de "AVISOS"
                        break

                # Se "AVISOS" foi encontrado, processa as linhas abaixo dele
                if aviso_index is not None and aviso_index < len(df):
                    df = df.iloc[aviso_index:].reset_index(drop=True)  # Pega apenas as linhas abaixo de "AVISOS"
                    df.columns = colunas[:len(df.columns)]  # Ajusta as colunas às esperadas
                    df = df.applymap(lambda x: limpar_quebras_de_linha(x) if isinstance(x, str) else x)

                    # Padroniza o layout para garantir colunas consistentes
                    df = padronizar_layout(df, colunas)

                    # Adiciona as linhas ao conjunto de dados completos
                    dados_completos.extend(df.values.tolist())

            # Verifica se dados foram extraídos
            if not dados_completos:
                messagebox.showwarning("Nenhum dado encontrado", "Nenhuma tabela de AVISOS foi encontrada no PDF.")
                return

            # Converte os dados em um DataFrame organizado com layout padronizado
            df_final = pd.DataFrame(dados_completos, columns=colunas)

            # Exibe os dados extraídos em uma janela de texto
            exibir_dados(df_final)

            # Solicita onde salvar o arquivo Excel
            excel_path = filedialog.asksaveasfilename(
                title="Salvar arquivo Excel como",
                defaultextension=".xlsx",
                filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
            )

            if excel_path:
                # Salva em uma única aba com colunas definidas
                df_final.to_excel(excel_path, sheet_name="Dados AVISOS", index=False)

                # Formatação adicional usando openpyxl
                workbook = load_workbook(excel_path)
                sheet = workbook["Dados AVISOS"]

                # Configuração de cabeçalhos: Negrito e preenchimento
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                for col in range(1, len(colunas) + 1):
                    cell = sheet[f"{get_column_letter(col)}1"]
                    cell.font = header_font
                    cell.fill = header_fill

                # Ajusta a largura das colunas automaticamente
                for col in sheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
                    sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length

                # Alinha o conteúdo das células ao centro
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

                # Adiciona filtros na linha de cabeçalho
                sheet.auto_filter.ref = sheet.dimensions

                # Salva o arquivo Excel formatado
                workbook.save(excel_path)
                messagebox.showinfo("Sucesso", f"Arquivo Excel salvo em:\n{excel_path}")
            else:
                messagebox.showwarning("Cancelado", "Operação de salvamento cancelada.")

        except Exception as e:
            # Mensagem de erro caso ocorra algum problema
            messagebox.showerror("Erro", f"Falha ao processar o PDF:\n{str(e)}")
    else:
        messagebox.showwarning("Cancelado", "Operação de carregamento cancelada.")

# Função para exibir dados em uma janela de texto
def exibir_dados(df):
    # Janela de exibição
    janela_exibir = tk.Toplevel(app)
    janela_exibir.title("Dados Extraídos")
    janela_exibir.geometry("800x400")

    # Caixa de texto com rolagem
    texto = scrolledtext.ScrolledText(janela_exibir, wrap=tk.WORD, width=100, height=20)
    texto.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    # Insere os dados no formato tabular
    texto.insert(tk.END, df.to_string(index=False))

    # Torna a caixa de texto apenas para leitura
    texto.configure(state='disabled')

# Configurações da interface Tkinter
app = tk.Tk()
app.title("Conversor PDF para Excel")
app.geometry("400x200")

# Botão para carregar PDF e converter
botao_carregar = tk.Button(app, text="Carregar PDF e Converter para Excel", command=carregar_pdf)
botao_carregar.pack(pady=20)

# Inicia a interface
app.mainloop()