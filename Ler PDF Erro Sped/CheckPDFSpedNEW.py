import tkinter as tk
from tkinter import filedialog, messagebox
import camelot
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# Função para remover quebras de linha dentro das células
def limpar_quebras_de_linha(texto):
    return texto.replace("\n", " ") if isinstance(texto, str) else texto

# Função para carregar PDF e salvar como Excel
def carregar_pdf():
    # Solicita ao usuário o arquivo PDF
    pdf_path = filedialog.askopenfilename(
        title="Selecione o arquivo PDF",
        filetypes=(("PDF Files", "*.pdf"), ("All Files", "*.*"))
    )
    if pdf_path:
        try:
            # Lê todas as tabelas do PDF com o layout padrão para todas as páginas
            tables = camelot.read_pdf(pdf_path, pages="all", flavor="stream", strip_text="\n")

            # Estrutura esperada das colunas
            colunas = [
                "Linha", "Posição", "Campo", "Mensagem", "Registro",
                "Conteúdo do Registro", "Conteúdo do Campo", "Valor Esperado"
            ]
            dados_completos = []

            # Processa cada tabela e organiza as colunas
            for table in tables:
                df = table.df
                # Aplica a função limpar_quebras_de_linha a cada célula do DataFrame
                df = df.applymap(lambda x: limpar_quebras_de_linha(x) if isinstance(x, str) else x)

                # Verifica e adapta cada linha para a estrutura de colunas correta
                for _, row in df.iterrows():
                    # Inicializa uma linha vazia com o número de colunas
                    linha_dados = [""] * len(colunas)

                    # Aloca cada campo à coluna correta, removendo quebras de linha
                    linha_dados[0] = row[0] if len(row) > 0 else ""  # Linha
                    linha_dados[1] = row[1] if len(row) > 1 else ""  # Posição
                    linha_dados[2] = row[2] if len(row) > 2 else ""  # Campo
                    linha_dados[3] = row[3] if len(row) > 3 else ""  # Mensagem
                    linha_dados[4] = row[4] if len(row) > 4 else ""  # Registro
                    linha_dados[5] = row[5] if len(row) > 5 else ""  # Conteúdo do Registro
                    linha_dados[6] = row[6] if len(row) > 6 else ""  # Conteúdo do Campo
                    linha_dados[7] = row[7] if len(row) > 7 else ""  # Valor Esperado

                    # Adiciona a linha estruturada aos dados completos
                    dados_completos.append(linha_dados)

            # Converte os dados em um DataFrame organizado
            df_final = pd.DataFrame(dados_completos, columns=colunas)

            # Solicita onde salvar o arquivo Excel
            excel_path = filedialog.asksaveasfilename(
                title="Salvar arquivo Excel como",
                defaultextension=".xlsx",
                filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
            )

            if excel_path:
                # Salva em uma única aba com colunas definidas
                df_final.to_excel(excel_path, sheet_name="Dados Extraídos", index=False)

                # Formatação adicional usando openpyxl
                workbook = load_workbook(excel_path)
                sheet = workbook["Dados Extraídos"]

                # Configuração de cabeçalhos: Negrito e preenchimento
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                for col in range(1, len(colunas) + 1):
                    cell = sheet[f"{get_column_letter(col)}1"]
                    cell.font = header_font
                    cell.fill = header_fill

                # Ajusta a largura das colunas automaticamente
                for col in sheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
                    sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length

                # Alinha o conteúdo das células ao centro
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")

                # Adiciona filtros na linha de cabeçalho
                sheet.auto_filter.ref = sheet.dimensions

                # Salva o arquivo Excel formatado
                workbook.save(excel_path)
                messagebox.showinfo("Sucesso", f"Arquivo Excel salvo em:\n{excel_path}")
            else:
                messagebox.showwarning("Cancelado", "Operação de salvamento cancelada.")

        except Exception as e:
            # Mensagem de erro caso ocorra algum problema
            messagebox.showerror("Erro", f"Falha ao processar o PDF:\n{str(e)}")
    else:
        messagebox.showwarning("Cancelado", "Operação de carregamento cancelada.")

# Configurações da interface Tkinter
app = tk.Tk()
app.title("Conversor PDF para Excel")
app.geometry("400x200")

# Botão para carregar PDF e converter
botao_carregar = tk.Button(app, text="Carregar PDF e Converter para Excel", command=carregar_pdf)
botao_carregar.pack(pady=20)

# Inicia a interface
app.mainloop()